"""Importador estructurado y reutilizable de instrumentos Excel de Portafolio."""

import hashlib
import json
from io import BytesIO
from pathlib import Path

from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import (
    CalculadoraInstrumento,
    CategoriaDocumento,
    Documento,
    ImportacionInstrumento,
    Instrumento,
    PreguntaInstrumento,
    RevisionInstrumento,
)


HOJAS_REQUERIDAS = {
    'INSTRUMENTO',
    'PREGUNTAS',
    'CALCULADORA_SISTEMA',
    'CASOS_PRUEBA',
}

TIPOS = {
    'escala',
    'si_no',
    'opcion_unica',
    'opcion_multiple',
    'texto_libre',
}


def _clave_normalizada(valor):
    return ''.join(
        caracter
        for caracter in str(valor or '').lower()
        if caracter.isalnum()
    )


def _valor_metadato(metadatos, *nombres):
    claves = {_clave_normalizada(nombre) for nombre in nombres}
    for nombre, valor in metadatos.items():
        if _clave_normalizada(nombre) in claves and valor not in (None, ''):
            return valor
    return None


def _metadatos_instrumento(filas):
    """Conserva los pares del Excel y normaliza instrucciones en texto libre."""
    metadatos = {}
    instrucciones = []
    leyendo_instrucciones = False

    for fila in filas:
        etiqueta = str(fila[0]).strip() if fila and fila[0] is not None else ''
        valor = fila[1] if len(fila) > 1 else None
        etiqueta_normalizada = _clave_normalizada(etiqueta)

        if etiqueta and valor not in (None, ''):
            metadatos[etiqueta] = valor
            leyendo_instrucciones = False
            continue
        if etiqueta and 'instruccion' in etiqueta_normalizada:
            metadatos[etiqueta] = valor
            leyendo_instrucciones = True
            continue
        if (
            etiqueta
            and valor in (None, '')
            and leyendo_instrucciones
            and not etiqueta_normalizada.startswith(('aviso', 'nota'))
        ):
            instrucciones.append(etiqueta)
            continue
        if etiqueta:
            metadatos[etiqueta] = valor
        leyendo_instrucciones = False

    if instrucciones:
        metadatos['instrucciones'] = '\n\n'.join(instrucciones)
    return metadatos


def _booleano(valor):
    if isinstance(valor, bool):
        return valor
    return str(valor or '').strip().lower() in {'1', 'true', 'si', 'sí'}


def _condicion_visibilidad(valor):
    if not isinstance(valor, str):
        return valor or None
    try:
        return json.loads(valor)
    except json.JSONDecodeError as error:
        raise ValidationError('PREGUNTAS contiene una condición de visibilidad inválida.') from error


def _campos_contexto_requeridos(metadatos, primera):
    campos = set()
    for nombre, valor in metadatos.items():
        clave = _clave_normalizada(nombre)
        if 'contexto' not in clave and 'requer' not in clave:
            continue
        valores = valor if isinstance(valor, (list, tuple, set)) else str(valor or '').split(',')
        campos.update(
            str(campo).strip().lower()
            for campo in valores
            if str(campo).strip().lower() in {'sexo', 'fecha_nacimiento', 'edad'}
        )
    if primera.get('edad_min') is not None or primera.get('edad_max') is not None:
        campos.add('fecha_nacimiento')
    return sorted(campos)


def _filas(hoja):
    return list(
        hoja.iter_rows(
            values_only=True,
        )
    )


def _tabla(filas, encabezado):
    for indice, fila in enumerate(filas):
        if any(
            str(valor or '').strip() == encabezado
            for valor in fila
        ):
            columnas = [
                str(valor or '').strip()
                for valor in fila
            ]

            return [
                {
                    columnas[i]: valor
                    for i, valor in enumerate(actual)
                    if i < len(columnas) and columnas[i]
                }
                for actual in filas[indice + 1:]
                if any(
                    valor is not None and str(valor).strip()
                    for valor in actual
                )
            ]

    return []


def leer_excel(origen):
    nombre = Path(getattr(origen, 'name', origen)).name
    try:
        if hasattr(origen, 'read'):
            origen.seek(0)
            contenido = origen.read()
            origen.seek(0)
        else:
            contenido = Path(origen).read_bytes()
        libro = load_workbook(
            BytesIO(contenido),
            read_only=True,
            data_only=False,
        )
    except Exception as error:
        raise ValidationError(
            f'{nombre}: no fue posible leer el archivo.'
        ) from error

    faltantes = HOJAS_REQUERIDAS - set(libro.sheetnames)

    if faltantes:
        libro.close()
        raise ValidationError(
            f'{nombre}: faltan hojas requeridas: '
            f'{", ".join(sorted(faltantes))}.'
        )

    instrumento_filas = _filas(
        libro['INSTRUMENTO']
    )

    valores = _metadatos_instrumento(instrumento_filas)

    preguntas = _tabla(
        _filas(libro['PREGUNTAS']),
        'instrumento_clave',
    )

    calculadora_filas = _filas(
        libro['CALCULADORA_SISTEMA']
    )

    calculadora = {
        str(f[0]).strip(): f[1]
        for f in calculadora_filas
        if (
            len(f) > 1
            and f[0]
            and str(f[0]).strip() not in {'Campo'}
        )
    }

    casos = _tabla(
        _filas(libro['CASOS_PRUEBA']),
        'Caso',
    )

    if not preguntas:
        libro.close()
        raise ValidationError(
            f'{nombre}: PREGUNTAS no contiene filas.'
        )

    if (
        not calculadora.get('clave_calculadora')
        or not calculadora.get('version_regla')
    ):
        libro.close()
        raise ValidationError(
            f'{nombre}: CALCULADORA_SISTEMA requiere '
            'clave_calculadora y version_regla.'
        )

    estructura_para_huella = {
        'instrumento': valores,
        'preguntas': preguntas,
        'calculadora': calculadora,
        'casos': casos,
    }
    huella = hashlib.sha256(
        json.dumps(
            estructura_para_huella,
            sort_keys=True,
            ensure_ascii=False,
            default=str,
        ).encode('utf-8'),
    ).hexdigest()

    datos = {
        'nombre_archivo': nombre,
        'contenido': contenido,
        'huella': huella,
        'instrumento': valores,
        'preguntas': preguntas,
        'calculadora': calculadora,
        'calculadora_filas': calculadora_filas,
        'casos': casos,
    }
    libro.close()
    return datos


def validar(datos):
    preguntas = datos['preguntas']
    claves = set()
    ordenes = set()

    for numero, pregunta in enumerate(
        preguntas,
        2,
    ):
        clave, orden, texto, tipo = (
            pregunta.get('pregunta_clave'),
            pregunta.get('orden'),
            pregunta.get('texto'),
            str(
                pregunta.get('tipo_respuesta') or ''
            ).strip(),
        )

        if not clave or clave in claves:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'clave duplicada o vacía.'
            )

        if not isinstance(orden, int) or orden in ordenes:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'orden duplicado o inválido.'
            )

        if not texto:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'texto vacío.'
            )

        if tipo not in TIPOS:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                f'tipo no soportado: {tipo}.'
            )

        try:
            opciones = json.loads(
                pregunta.get('opciones_json') or 'null'
            )
        except json.JSONDecodeError as error:
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'opciones_json inválido.'
            ) from error

        if opciones is not None and (
            not isinstance(opciones, list)
            or any(
                not isinstance(o, dict)
                or 'valor' not in o
                or 'etiqueta' not in o
                for o in opciones
            )
        ):
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS fila {numero}, '
                'opciones_json inválido.'
            )

        claves.add(clave)
        ordenes.add(orden)

    primera = preguntas[0]

    for campo in (
        'instrumento_clave',
        'instrumento_nombre',
        'version',
    ):
        if not primera.get(campo):
            raise ValidationError(
                f'{datos["ruta"].name}: PREGUNTAS requiere {campo}.'
            )

    if str(datos['calculadora']['version_regla']) != str(primera['version']):
        raise ValidationError(
            f'{datos["ruta"].name}: version_regla debe coincidir con la versión del instrumento.'
        )

    return datos


def _estado(valor):
    texto = str(valor or '').upper()

    if 'BLOQUE' in texto:
        return CalculadoraInstrumento.Estado.BLOQUEADA

    if 'NO ACTIVA' in texto or 'DIAGN' in texto:
        return CalculadoraInstrumento.Estado.NO_DIAGNOSTICA

    if 'ORIENTAT' in texto:
        return CalculadoraInstrumento.Estado.ORIENTATIVA

    return CalculadoraInstrumento.Estado.ACTIVA


def _estado_calculadora_por_instrumento(clave, calculadora):
    """Conserva la interpretación orientativa de las variantes autorizadas."""
    estado = _estado(
        calculadora.get('estado_calculadora')
        or calculadora.get('estado_puntaje')
        or calculadora.get('estado_interpretacion')
    )
    if clave in {
        'dass-21-adolescentes',
        'rse-autoestima',
    } and _estado(calculadora.get('estado_interpretacion')) == (
        CalculadoraInstrumento.Estado.ORIENTATIVA
    ):
        return CalculadoraInstrumento.Estado.ORIENTATIVA
    return estado


def _definicion(datos):
    clave = datos['preguntas'][0]['instrumento_clave']
    clase = 'conteo'

    if clave.startswith('dass-21'):
        clase = 'subescalas_multiplicadas'
    elif clave == 'rse-autoestima':
        clase = 'suma_recodificada'
    elif 'plutchik' in clave:
        clase = 'suma_criticos'

    return {
        'algoritmo': clase,
        'campos': datos['calculadora'],
        'tablas': datos['calculadora_filas'],
        'casos_prueba': datos['casos'],
        'requiere_respuestas_completas': str(
            datos['calculadora'].get(
                'requiere_respuestas_completas',
                '',
            )
        ).lower() in {
            'true',
            'sí',
            'si',
        },
        'sin_acciones_automaticas': True,
    }


def ejecutar_calculadora(
    definicion,
    respuestas,
    edad=None,
):
    if (
        definicion.get('estado')
        == CalculadoraInstrumento.Estado.BLOQUEADA
    ):
        return {
            'bloqueada': True,
        }

    algoritmo = definicion['algoritmo']

    valores = {
        int(k): int(v)
        for k, v in respuestas.items()
        if v not in (
            None,
            '',
        )
    }

    if algoritmo == 'subescalas_multiplicadas':
        grupos = {
            'Depresión': [
                3,
                5,
                10,
                13,
                16,
                17,
                21,
            ],
            'Ansiedad': [
                2,
                4,
                7,
                9,
                15,
                19,
                20,
            ],
            'Estrés': [
                1,
                6,
                8,
                11,
                12,
                14,
                18,
            ],
        }

        detalle = {
            n: sum(
                valores.get(i, 0)
                for i in items
            ) * 2
            for n, items in grupos.items()
        }

        return {
            'puntaje_total': sum(detalle.values()),
            'detalle': detalle,
        }

    if algoritmo == 'suma_recodificada':
        inversos = {
            2,
            5,
            8,
            9,
            10,
        }

        total = sum(
            5 - v if i in inversos else v
            for i, v in valores.items()
        )

        nivel = (
            'Autoestima elevada'
            if total >= 30
            else (
                'Autoestima media'
                if total >= 26
                else (
                    'SIN RANGO DEFINIDO'
                    if total == 25
                    else 'Autoestima baja'
                )
            )
        )

        return {
            'puntaje_total': total,
            'interpretacion': nivel,
            'detalle': {},
        }

    if algoritmo == 'suma_criticos':
        total = sum(valores.values())

        return {
            'puntaje_total': total,
            'interpretacion': (
                'Presencia de riesgo; evaluación inmediata'
                if total >= 6
                else 'No se detecta riesgo mediante este tamizaje'
            ),
            'focos': [
                i
                for i in (
                    13,
                    14,
                    15,
                )
                if valores.get(i) == 1
            ],
        }

    return {
        'puntaje_total': sum(valores.values()),
        'detalle': {
            'conteos': sum(valores.values()),
            'edad': edad,
        },
    }


def importar(origen, dry_run=False, nombre_archivo=None, cargado_por=None):
    datos = validar(
        leer_excel(origen)
    )

    primera = datos['preguntas'][0]
    clave = primera['instrumento_clave']
    version = str(primera['version'])
    nombre_archivo = Path(nombre_archivo or datos['nombre_archivo']).name
    definicion = _definicion(datos)

    estado = _estado_calculadora_por_instrumento(
        clave,
        datos['calculadora'],
    )

    reporte = {
        'archivo': nombre_archivo,
        'clave': clave,
        'version': version,
        'preguntas': len(datos['preguntas']),
        'casos': len(datos['casos']),
        'estado_calculadora': estado,
        'decision': 'importable',
        'dry_run': dry_run,
    }

    try:
        existente = Instrumento.objects.get(clave=clave)
    except Instrumento.DoesNotExist:
        existente = None

    if (
        existente
        and hasattr(existente, 'importacion')
        and existente.importacion.huella_contenido == datos['huella']
    ):
        reporte['decision'] = 'sin cambios'
        return reporte

    if existente and (
        existente.aplicaciones_intera.exists()
        or existente.preguntas.filter(
            respuestas_intera__isnull=False
        ).exists()
    ):
        raise ValidationError(
            f'{clave}: instrumento utilizado; no se sobrescribe.'
        )

    if dry_run:
        return reporte

    with transaction.atomic():
        categoria, _ = CategoriaDocumento.objects.get_or_create(
            nombre='Instrumento'
        )

        documentos = list(
            Documento.objects.filter(
                importaciones_instrumento__huella_contenido=datos['huella'],
            ).distinct()
        )
        if len(documentos) > 1:
            raise ValidationError(
                f'{clave}: existen varios Documentos con la misma huella; requiere revisión administrativa.'
            )
        documento = documentos[0] if documentos else None

        if not documento:
            documento = Documento(
                nombre=Path(nombre_archivo).stem,
                categoria=categoria,
                version=version,
                cargado_por=cargado_por,
                descripcion=(
                    'Documento origen importado por Portafolio.'
                ),
            )

            documento.archivo.save(
                nombre_archivo,
                ContentFile(datos['contenido']),
                save=False,
            )

            documento.save()

        instrumento = existente or Instrumento(
            clave=clave
        )

        instrumento.nombre = primera['instrumento_nombre']
        instrumento.version = version
        instrumento.documento_origen = documento
        instrumento.descripcion = str(
            _valor_metadato(datos['instrumento'], 'descripcion') or ''
        )
        instrumento.instrucciones = str(
            _valor_metadato(
                datos['instrumento'],
                'instrucciones',
                'instrucciones para la aplicacion',
                'instrucciones para la persona evaluada',
            ) or ''
        )
        instrumento.activo = True
        instrumento.full_clean()
        instrumento.save()

        ImportacionInstrumento.objects.update_or_create(
            instrumento=instrumento,
            defaults={
                'documento': documento,
                'huella_contenido': datos['huella'],
                'metadatos': {
                    **datos['instrumento'],
                    'variante': (
                        primera.get('variante')
                        or _valor_metadato(datos['instrumento'], 'variante')
                    ),
                    'poblacion': (
                        primera.get('poblacion')
                        or _valor_metadato(
                            datos['instrumento'],
                            'poblacion',
                            'poblacion objetivo',
                        )
                    ),
                    'edad_min': primera.get('edad_min'),
                    'edad_max': primera.get('edad_max'),
                    'campos_contexto_requeridos': _campos_contexto_requeridos(
                        datos['instrumento'],
                        primera,
                    ),
                },
            },
        )

        instrumento.preguntas.all().delete()

        PreguntaInstrumento.objects.bulk_create(
            [
                PreguntaInstrumento(
                    instrumento=instrumento,
                    orden=p['orden'],
                    clave=p['pregunta_clave'],
                    texto=p['texto'],
                    tipo_respuesta=p['tipo_respuesta'],
                    opciones=json.loads(
                        p.get('opciones_json') or 'null'
                    ),
                    requerida=_booleano(p.get('requerida')),
                    condicion_visibilidad=(
                        _condicion_visibilidad(p.get('visibilidad'))
                    ),
                )
                for p in datos['preguntas']
            ]
        )

        RevisionInstrumento.objects.update_or_create(
            instrumento=instrumento,
            version=version,
            defaults={
                'estructura': {
                    'metadatos': instrumento.importacion.metadatos,
                    'preguntas': list(
                        instrumento.preguntas.values(
                            'orden',
                            'clave',
                            'texto',
                            'tipo_respuesta',
                            'opciones',
                            'requerida',
                        )
                    ),
                },
            },
        )

        calculadora, _ = CalculadoraInstrumento.objects.update_or_create(
            instrumento=instrumento,
            clave=datos['calculadora']['clave_calculadora'],
            version_regla=str(
                datos['calculadora']['version_regla']
            ),
            defaults={
                'estado': estado,
                'definicion': {
                    **definicion,
                    'estado': estado,
                },
                'huella_contenido': datos['huella'],
            },
        )
        CalculadoraInstrumento.objects.filter(
            instrumento=instrumento,
            version_regla=version,
        ).exclude(pk=calculadora.pk).delete()

    return reporte


def importar_archivo_subido(archivo, cargado_por=None):
    """Entrega un archivo web al importador estructurado sin otro parser."""
    nombre_archivo = Path(archivo.name).name
    return importar(
        archivo,
        nombre_archivo=nombre_archivo,
        cargado_por=cargado_por,
    )
