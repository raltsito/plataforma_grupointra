from io import BytesIO
from types import SimpleNamespace
import json
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.db.migrations.executor import MigrationExecutor
from django.test import TestCase, TransactionTestCase
from django.test.utils import override_settings
from django.urls import reverse
from openpyxl import Workbook

from .models import (
    CalculadoraInstrumento,
    CategoriaDocumento,
    Documento,
    ImportacionInstrumento,
    Instrumento,
    PreguntaInstrumento,
    RevisionInstrumento,
)
from .services_entrevista import (
    CLAVE_ENTREVISTA,
    cargar_plantilla_entrevista_1a1,
)
from .services_importacion_instrumentos import (
    _estado_calculadora_por_instrumento,
    ejecutar_calculadora,
)
from .services_calificacion import (
    ADVERTENCIA_CORTA_RESULTADO_ORIENTATIVO,
    ADVERTENCIA_RESULTADO_ORIENTATIVO,
    calcular_resultado,
    campos_contexto_requeridos,
    edad_cumplida,
    obtener_revision_calculadora,
    validar_variante_por_edad,
)


@override_settings(
    STORAGES={
        'default': {
            'BACKEND': (
                'django.core.files.storage.'
                'FileSystemStorage'
            ),
        },
        'staticfiles': {
            'BACKEND': (
                'django.contrib.staticfiles.storage.'
                'StaticFilesStorage'
            ),
        },
    }
)
class ImportacionInstrumentosTests(TestCase):
    def _excel(self):
        libro = Workbook()
        hoja = libro.active

        hoja.append(
            [
                'Nombre completo|name-1',
                '1. ¿Primera pregunta?|radio-1',
                '2. Explica tu respuesta|textarea-2',
            ]
        )

        contenido = BytesIO()
        libro.save(contenido)

        return SimpleUploadedFile(
            'instrumento.xlsx',
            contenido.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

    def test_importa_preguntas_desde_el_documento_origen_de_portafolio(
        self,
    ):
        documento = Documento.objects.create(
            nombre='SCID fuente',
            archivo=self._excel(),
        )

        instrumento = Instrumento.objects.create(
            nombre='SCID',
            clave='scid',
            documento_origen=documento,
        )

        usuario = User.objects.create_user(
            username='portafolio',
            password='secreto',
        )

        grupo, _ = Group.objects.get_or_create(
            name='Certificación'
        )

        usuario.groups.add(grupo)
        self.client.force_login(usuario)

        respuesta = self.client.post(
            reverse(
                'portafolio:importar_preguntas',
                args=[instrumento.id],
            )
        )

        self.assertRedirects(
            respuesta,
            reverse('portafolio:instrumentos'),
        )

        preguntas = list(
            instrumento.preguntas.order_by('orden')
        )

        self.assertEqual(
            len(preguntas),
            2,
        )

        self.assertEqual(
            preguntas[0].opciones,
            [
                {
                    'valor': '1',
                    'etiqueta': 'Sí',
                },
                {
                    'valor': '0',
                    'etiqueta': 'No',
                },
            ],
        )

        self.assertEqual(
            preguntas[1].tipo_respuesta,
            'texto_libre',
        )


@override_settings(
    STORAGES={
        'default': {
            'BACKEND': 'django.core.files.storage.FileSystemStorage',
        },
        'staticfiles': {
            'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage',
        },
    }
)
class ImportacionInstrumentoWebTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(
            username='importador',
            password='secreto',
        )
        grupo, _ = Group.objects.get_or_create(name='Certificación')
        self.usuario.groups.add(grupo)
        self.client.force_login(self.usuario)
        self.documentos_iniciales = Documento.objects.count()
        self.instrumentos_iniciales = Instrumento.objects.count()

    def _excel_estructurado(self):
        libro = Workbook()
        instrumento = libro.active
        instrumento.title = 'INSTRUMENTO'
        instrumento.append(['Campo', 'Valor'])
        instrumento.append(['Población objetivo', 'Adolescentes'])
        instrumento.append(['Instrucciones', None])
        instrumento.append(['Lee cada frase y elige la respuesta más adecuada.', None])

        preguntas = libro.create_sheet('PREGUNTAS')
        preguntas.append([
            'instrumento_clave', 'instrumento_nombre', 'variante',
            'version', 'poblacion', 'edad_min', 'edad_max', 'orden',
            'pregunta_clave', 'texto', 'tipo_respuesta', 'opciones_json',
            'requerida', 'visibilidad',
        ])
        preguntas.append([
            'instrumento-web', 'Instrumento web', 'Adolescentes', '1.0',
            'Adolescentes', 14, 20, 1, 'P-01', 'Primera pregunta',
            'si_no', '[{"valor":"si","etiqueta":"Sí"},{"valor":"no","etiqueta":"No"}]',
            True, None,
        ])
        preguntas.append([
            'instrumento-web', 'Instrumento web', 'Adolescentes', '1.0',
            'Adolescentes', 14, 20, 2, 'P-02', 'Explica tu respuesta',
            'texto_libre', None, False,
            '{"pregunta_clave":"P-01","operador":"igual","valor":"si"}',
        ])

        calculadora = libro.create_sheet('CALCULADORA_SISTEMA')
        calculadora.append(['Campo', 'Valor'])
        calculadora.append(['clave_calculadora', 'calc-instrumento-web-v1'])
        calculadora.append(['version_regla', '1.0'])
        calculadora.append(['estado_calculadora', 'ORIENTATIVA'])
        calculadora.append(['requiere_respuestas_completas', True])

        casos = libro.create_sheet('CASOS_PRUEBA')
        casos.append(['Caso'])

        contenido = BytesIO()
        libro.save(contenido)
        return SimpleUploadedFile(
            'instrumento-web.xlsx',
            contenido.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

    def test_importa_desde_instrumentos_toda_la_estructura(self):
        respuesta = self.client.post(
            reverse('portafolio:instrumentos'),
            {'accion': 'importar', 'archivo': self._excel_estructurado()},
        )

        self.assertRedirects(respuesta, reverse('portafolio:instrumentos'))
        instrumento = Instrumento.objects.get(clave='instrumento-web')
        self.assertTrue(instrumento.activo)
        self.assertEqual(instrumento.preguntas.count(), 2)
        self.assertIn('Lee cada frase', instrumento.instrucciones)
        self.assertEqual(instrumento.documento_origen.nombre, 'instrumento-web')
        self.assertEqual(instrumento.documento_origen.cargado_por, self.usuario)
        self.assertTrue(
            instrumento.documento_origen.archivo.storage.exists(
                instrumento.documento_origen.archivo.name,
            ),
        )
        self.assertEqual(
            instrumento.importacion.metadatos['campos_contexto_requeridos'],
            ['fecha_nacimiento'],
        )
        self.assertEqual(
            campos_contexto_requeridos(instrumento),
            {'fecha_nacimiento'},
        )
        self.assertEqual(
            CalculadoraInstrumento.objects.get(
                instrumento=instrumento,
            ).estado,
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        pregunta = PreguntaInstrumento.objects.get(instrumento=instrumento, clave='P-02')
        self.assertFalse(pregunta.requerida)
        self.assertEqual(pregunta.condicion_visibilidad['pregunta_clave'], 'P-01')
        from apps.certificacion_intera.views import _instrumentos_para_bateria
        _, seleccionables = _instrumentos_para_bateria()
        self.assertIn(instrumento, seleccionables)

    def test_reimportar_el_mismo_archivo_no_duplica_registros(self):
        for _ in range(2):
            self.client.post(
                reverse('portafolio:instrumentos'),
                {'accion': 'importar', 'archivo': self._excel_estructurado()},
            )

        instrumento = Instrumento.objects.get(clave='instrumento-web')
        self.assertEqual(Documento.objects.filter(nombre='instrumento-web').count(), 1)
        self.assertEqual(Instrumento.objects.filter(clave='instrumento-web').count(), 1)
        self.assertEqual(ImportacionInstrumento.objects.filter(instrumento=instrumento).count(), 1)
        self.assertEqual(PreguntaInstrumento.objects.filter(instrumento=instrumento).count(), 2)
        self.assertEqual(CalculadoraInstrumento.objects.filter(instrumento=instrumento).count(), 1)

    def test_excel_invalido_muestra_error_y_no_crea_registros(self):
        archivo = SimpleUploadedFile(
            'invalido.xlsx',
            b'no es un libro de Excel',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        respuesta = self.client.post(
            reverse('portafolio:instrumentos'),
            {'accion': 'importar', 'archivo': archivo},
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'no fue posible leer el archivo')
        self.assertEqual(Documento.objects.count(), self.documentos_iniciales)
        self.assertEqual(Instrumento.objects.count(), self.instrumentos_iniciales)

    def test_conserva_el_alta_manual_de_instrumentos(self):
        respuesta = self.client.post(
            reverse('portafolio:instrumentos'),
            {
                'nombre': 'Instrumento manual',
                'clave': 'instrumento-manual',
                'activo': 'on',
            },
        )

        self.assertRedirects(respuesta, reverse('portafolio:instrumentos'))
        self.assertTrue(
            Instrumento.objects.filter(
                clave='instrumento-manual',
                activo=True,
            ).exists()
        )

    def test_catalogo_solo_presenta_el_importador_estructurado(self):
        respuesta = self.client.get(reverse('portafolio:instrumentos'))

        self.assertContains(respuesta, 'Subir archivo')
        self.assertContains(respuesta, 'Subir archivo e importar instrumento')
        self.assertNotContains(respuesta, 'Importar preguntas del Excel')
        self.assertNotContains(
            respuesta,
            reverse('portafolio:importar_preguntas', args=[1]),
        )


@override_settings(
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    },
)
class DescargaDocumentosTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(
            username='descarga-documentos',
            password='secreto',
        )
        self.usuario.groups.add(
            Group.objects.get_or_create(name='Certificación')[0],
        )
        self.client.force_login(self.usuario)

    def test_descarga_protegida_usa_storage_y_nombre_original(self):
        documento = Documento.objects.create(
            nombre='Documento descargable',
            archivo=SimpleUploadedFile('reporte seguro.xlsx', b'contenido'),
        )

        respuesta = self.client.get(
            reverse('portafolio:documento_descargar', args=[documento.id]),
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertIn('attachment;', respuesta['Content-Disposition'])
        self.assertIn('reporte_seguro', respuesta['Content-Disposition'])
        self.assertNotIn('/media/', respuesta['Content-Disposition'])
        respuesta.close()
        documento.archivo.delete(save=False)

    def test_archivo_faltante_devuelve_404_controlado(self):
        documento = Documento.objects.create(
            nombre='Documento perdido',
            archivo=SimpleUploadedFile('perdido.xlsx', b'contenido'),
        )
        documento.archivo.storage.delete(documento.archivo.name)

        respuesta = self.client.get(
            reverse('portafolio:documento_descargar', args=[documento.id]),
        )

        self.assertEqual(respuesta.status_code, 404)
        self.assertContains(
            respuesta,
            'se encuentra actualmente en el almacenamiento configurado',
            status_code=404,
        )

    def test_documentos_enlaza_la_vista_protegida_y_no_media_url(self):
        documento = Documento.objects.create(
            nombre='Documento listado',
            archivo=SimpleUploadedFile('listado.xlsx', b'contenido'),
        )

        respuesta = self.client.get(reverse('portafolio:documentos'))

        self.assertContains(
            respuesta,
            reverse('portafolio:documento_descargar', args=[documento.id]),
        )
        self.assertNotContains(respuesta, documento.archivo.url)
        documento.archivo.delete(save=False)


@override_settings(
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    }
)
class EliminacionPortafolioTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(username='eliminador', password='secreto')
        self.usuario.groups.add(Group.objects.get_or_create(name='Certificación')[0])
        self.client.force_login(self.usuario)
        self.categoria = CategoriaDocumento.objects.get_or_create(nombre='Instrumento')[0]

    def _documento(self, nombre='Documento temporal'):
        return Documento.objects.create(
            nombre=nombre,
            categoria=self.categoria,
            archivo=SimpleUploadedFile('temporal.xlsx', b'contenido'),
        )

    def test_documento_huerfano_se_elimina_solo_por_post(self):
        documento = self._documento()
        almacenamiento = documento.archivo.storage
        nombre_archivo = documento.archivo.name
        url = reverse('portafolio:eliminar', args=['documento', documento.id])

        self.assertEqual(self.client.get(url).status_code, 200)
        self.assertTrue(Documento.objects.filter(pk=documento.pk).exists())
        with self.captureOnCommitCallbacks(execute=True):
            respuesta = self.client.post(url)

        self.assertRedirects(respuesta, reverse('portafolio:documentos'))
        self.assertFalse(Documento.objects.filter(pk=documento.pk).exists())
        self.assertFalse(almacenamiento.exists(nombre_archivo))

    def test_documento_usado_se_protege_sin_error_servidor(self):
        documento = self._documento()
        Instrumento.objects.create(
            nombre='Instrumento protegido',
            clave='instrumento-protegido',
            documento_origen=documento,
        )
        url = reverse('portafolio:eliminar', args=['documento', documento.id])

        respuesta = self.client.post(url, follow=True)

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'no puede eliminarse')
        self.assertTrue(Documento.objects.filter(pk=documento.pk).exists())

    def test_instrumento_sin_historial_se_elimina_con_sus_dependencias(self):
        documento = self._documento('Origen importado')
        instrumento = Instrumento.objects.create(
            nombre='Instrumento temporal',
            clave='instrumento-temporal',
            documento_origen=documento,
        )
        ImportacionInstrumento.objects.create(
            instrumento=instrumento,
            documento=documento,
            huella_contenido='a' * 64,
        )
        PreguntaInstrumento.objects.create(instrumento=instrumento, orden=1, texto='Pregunta')
        RevisionInstrumento.objects.create(instrumento=instrumento, version='1.0', estructura={})
        CalculadoraInstrumento.objects.create(
            instrumento=instrumento,
            clave='calc-temporal',
            version_regla='1.0',
            estado=CalculadoraInstrumento.Estado.ACTIVA,
            huella_contenido='b' * 64,
        )

        respuesta = self.client.post(
            reverse('portafolio:eliminar', args=['instrumento', instrumento.id]),
        )

        self.assertRedirects(respuesta, reverse('portafolio:instrumentos'))
        self.assertFalse(Instrumento.objects.filter(pk=instrumento.pk).exists())
        self.assertFalse(Documento.objects.filter(pk=documento.pk).exists())

    def test_instrumento_configurado_en_intera_no_se_elimina(self):
        from apps.certificacion_intera.models import (
            ConfiguracionInstrumento,
            Escuela,
            ProcesoCertificacion,
        )

        instrumento = Instrumento.objects.create(
            nombre='Instrumento con historial',
            clave='instrumento-con-historial',
        )
        escuela = Escuela.objects.create(
            nombre='Escuela', director='Dirección', cantidad_total_alumnos=1,
            estado='Estado', municipio='Municipio',
        )
        proceso = ProcesoCertificacion.objects.create(
            escuela=escuela,
            ciclo_escolar='2026-2027',
            fecha_inicio=date.today(),
        )
        ConfiguracionInstrumento.objects.create(proceso=proceso, instrumento=instrumento)

        respuesta = self.client.post(
            reverse('portafolio:eliminar', args=['instrumento', instrumento.id]),
            follow=True,
        )

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'no puede eliminarse')
        self.assertTrue(Instrumento.objects.filter(pk=instrumento.pk).exists())


class PlantillaEntrevistaUnoAUnoTests(TestCase):
    def test_plantilla_es_idempotente_y_publica_veinticuatro_preguntas(
        self,
    ):
        instrumento, _ = cargar_plantilla_entrevista_1a1()

        cargar_plantilla_entrevista_1a1()

        self.assertEqual(
            instrumento.clave,
            CLAVE_ENTREVISTA,
        )

        self.assertEqual(
            instrumento.preguntas.count(),
            24,
        )

        self.assertEqual(
            RevisionInstrumento.objects.filter(
                instrumento=instrumento
            ).count(),
            1,
        )


class CalculadorasEstructuradasTests(TestCase):
    def test_dass_suma_subescalas_y_multiplica(self):
        resultado = ejecutar_calculadora(
            {
                'algoritmo': 'subescalas_multiplicadas',
            },
            {
                i: 1
                for i in range(1, 22)
            },
        )

        self.assertEqual(
            resultado['detalle'],
            {
                'Depresión': 14,
                'Ansiedad': 14,
                'Estrés': 14,
            },
        )

    def test_rosenberg_recodifica_y_no_inventa_rango_veinticinco(
        self,
    ):
        resultado = ejecutar_calculadora(
            {
                'algoritmo': 'suma_recodificada',
            },
            {
                1: 4,
                2: 4,
                3: 4,
                4: 4,
                5: 4,
                6: 4,
                7: 4,
                8: 4,
                9: 4,
                10: 4,
            },
        )

        self.assertEqual(
            resultado['puntaje_total'],
            25,
        )

        self.assertEqual(
            resultado['interpretacion'],
            'SIN RANGO DEFINIDO',
        )

    def test_plutchik_bloqueada_no_se_ejecuta(self):
        self.assertEqual(
            ejecutar_calculadora(
                {
                    'estado': 'bloqueada',
                },
                {
                    1: 1,
                },
            ),
            {
                'bloqueada': True,
            },
        )


class RevisionCalculadoraTests(TestCase):
    def _instrumento_con_calculadora(self, estado):
        claves = {
            CalculadoraInstrumento.Estado.ACTIVA: 'dass21',
            CalculadoraInstrumento.Estado.ORIENTATIVA: 'dass_21',
            CalculadoraInstrumento.Estado.BLOQUEADA: 'dass-21',
            CalculadoraInstrumento.Estado.NO_DIAGNOSTICA: 'dass-21-r',
        }
        instrumento = Instrumento.objects.create(
            nombre=f'DASS {estado}',
            clave=claves[estado],
            version='1.0',
        )
        CalculadoraInstrumento.objects.create(
            instrumento=instrumento,
            clave=f'calculadora-{estado}',
            version_regla='1.0',
            estado=estado,
            definicion={},
            huella_contenido=f'huella-{estado}',
        )
        return instrumento

    def _respuestas_dass(self):
        return [
            SimpleNamespace(
                pregunta=SimpleNamespace(orden=orden),
                valor_numerico=1,
            )
            for orden in range(1, 22)
        ]

    def test_activa_ejecuta_sin_advertencia_orientativa(self):
        instrumento = self._instrumento_con_calculadora(
            CalculadoraInstrumento.Estado.ACTIVA
        )
        resultado = calcular_resultado(
            instrumento,
            self._respuestas_dass(),
        )

        self.assertTrue(resultado['revision_calculadora'])
        self.assertFalse(resultado['requiere_revision_profesional'])
        self.assertEqual(resultado['advertencia_larga'], '')
        self.assertEqual(resultado['advertencia_corta'], '')

    def test_orientativa_ejecuta_y_agrega_metadatos_sin_alterar_calculo(self):
        activa = self._instrumento_con_calculadora(
            CalculadoraInstrumento.Estado.ACTIVA
        )
        orientativa = self._instrumento_con_calculadora(
            CalculadoraInstrumento.Estado.ORIENTATIVA
        )

        resultado_activo = calcular_resultado(
            activa,
            self._respuestas_dass(),
        )
        resultado_orientativo = calcular_resultado(
            orientativa,
            self._respuestas_dass(),
        )

        self.assertTrue(resultado_orientativo['requiere_revision_profesional'])
        self.assertEqual(
            resultado_orientativo['advertencia_larga'],
            ADVERTENCIA_RESULTADO_ORIENTATIVO,
        )
        self.assertEqual(
            resultado_orientativo['advertencia_corta'],
            ADVERTENCIA_CORTA_RESULTADO_ORIENTATIVO,
        )
        self.assertEqual(
            resultado_orientativo['puntaje_total'],
            resultado_activo['puntaje_total'],
        )
        self.assertEqual(
            resultado_orientativo['interpretacion'],
            resultado_activo['interpretacion'],
        )
        detalle_orientativo = dict(resultado_orientativo['detalle'])
        detalle_activo = dict(resultado_activo['detalle'])
        detalle_orientativo.pop('trazabilidad_calculadora')
        detalle_activo.pop('trazabilidad_calculadora')
        self.assertEqual(detalle_orientativo, detalle_activo)

    def test_bloqueada_y_no_diagnostica_conservan_su_revision(self):
        bloqueada = self._instrumento_con_calculadora(
            CalculadoraInstrumento.Estado.BLOQUEADA
        )
        no_diagnostica = self._instrumento_con_calculadora(
            CalculadoraInstrumento.Estado.NO_DIAGNOSTICA
        )

        self.assertFalse(obtener_revision_calculadora(bloqueada)['puede_ejecutarse'])
        self.assertEqual(
            obtener_revision_calculadora(no_diagnostica)['estado_revision'],
            CalculadoraInstrumento.Estado.NO_DIAGNOSTICA,
        )

    def test_ausencia_o_ambiguedad_no_se_resuelve_con_first(self):
        sin_calculadora = Instrumento.objects.create(
            nombre='Sin calculadora',
            clave='sin-calculadora',
            version='1.0',
        )
        self.assertEqual(
            obtener_revision_calculadora(sin_calculadora)['seleccion'],
            'ausente',
        )

        CalculadoraInstrumento.objects.create(
            instrumento=sin_calculadora,
            clave='regla-a',
            version_regla='1.0',
            estado=CalculadoraInstrumento.Estado.ACTIVA,
            definicion={},
            huella_contenido='a',
        )
        CalculadoraInstrumento.objects.create(
            instrumento=sin_calculadora,
            clave='regla-b',
            version_regla='1.0',
            estado=CalculadoraInstrumento.Estado.ORIENTATIVA,
            definicion={},
            huella_contenido='b',
        )
        self.assertEqual(
            obtener_revision_calculadora(sin_calculadora)['seleccion'],
            'ambigua',
        )


class CalculadorasAdolescentesOrientativasTests(TestCase):
    def _instrumento(self, clave, estado):
        instrumento = Instrumento.objects.create(
            nombre=clave,
            clave=clave,
            version='1.0',
        )
        CalculadoraInstrumento.objects.create(
            instrumento=instrumento,
            clave=f'calc-{clave}-v1',
            version_regla='1.0',
            estado=estado,
            definicion={'estado': estado},
            huella_contenido=f'huella-{clave}',
        )
        return instrumento

    def test_dass_adolescente_usa_funcion_propia_y_conserva_puntajes(self):
        instrumento = self._instrumento(
            'dass-21-adolescentes',
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        respuestas = [
            SimpleNamespace(
                pregunta=SimpleNamespace(orden=orden, opciones=[]),
                valor_numerico=Decimal('1'),
            )
            for orden in range(1, 22)
        ]

        resultado = calcular_resultado(
            instrumento,
            respuestas,
            {
                'fecha_nacimiento': date(2010, 8, 6),
                'fecha_aplicacion': date(2026, 8, 6),
            },
        )

        self.assertEqual(resultado['detalle']['Depresión'], {
            'puntaje_bruto': 7,
            'puntaje_multiplicado': 14,
        })
        self.assertEqual(resultado['detalle']['Ansiedad']['puntaje_multiplicado'], 14)
        self.assertEqual(resultado['detalle']['Estrés']['puntaje_multiplicado'], 14)
        self.assertEqual(resultado['puntaje_total'], Decimal('42'))
        self.assertTrue(resultado['requiere_revision_profesional'])
        self.assertIn('referencias de la versión adulta', resultado['interpretacion'])
        self.assertNotIn('Normal', resultado['interpretacion'])
        json.dumps(resultado['detalle'])

    def test_rosenberg_invierte_conforme_a_las_opciones_importadas(self):
        instrumento = self._instrumento(
            'rse-autoestima',
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        opciones = [
            {'valor': str(valor), 'etiqueta': str(valor)}
            for valor in range(1, 5)
        ]
        respuestas = [
            SimpleNamespace(
                pregunta=SimpleNamespace(orden=orden, opciones=opciones),
                valor_numerico=Decimal('1'),
            )
            for orden in range(1, 11)
        ]

        resultado = calcular_resultado(
            instrumento,
            respuestas,
            {
                'fecha_nacimiento': date(2010, 8, 6),
                'fecha_aplicacion': date(2026, 8, 6),
            },
        )

        self.assertEqual(resultado['detalle']['puntaje_directos'], 5)
        self.assertEqual(resultado['detalle']['puntaje_inversos'], 20)
        self.assertEqual(resultado['puntaje_total'], Decimal('25'))
        self.assertEqual(
            resultado['detalle']['reactivos_inversos'][2]['rango_opciones'],
            [1, 4],
        )
        self.assertTrue(resultado['requiere_revision_profesional'])
        self.assertIn('no constituye un baremo adolescente validado', resultado['interpretacion'])
        json.dumps(resultado['detalle'])

    def test_importacion_preserva_estado_orientativo_solo_en_variantes_autorizadas(self):
        calculadora = {
            'estado_puntaje': 'ACTIVA',
            'estado_interpretacion': 'ORIENTATIVA',
        }
        self.assertEqual(
            _estado_calculadora_por_instrumento('dass-21-adolescentes', calculadora),
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        self.assertEqual(
            _estado_calculadora_por_instrumento('rse-autoestima', calculadora),
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        self.assertEqual(
            _estado_calculadora_por_instrumento('dass-21', calculadora),
            CalculadoraInstrumento.Estado.ACTIVA,
        )

    def test_limites_de_edad_y_cumpleanos_no_se_solapan(self):
        instrumento = self._instrumento(
            'dass-21-adolescentes',
            CalculadoraInstrumento.Estado.ORIENTATIVA,
        )
        fecha_aplicacion = date(2026, 8, 6)
        self.assertEqual(edad_cumplida(date(2007, 8, 7), fecha_aplicacion), 18)
        self.assertEqual(edad_cumplida(date(2007, 8, 6), fecha_aplicacion), 19)
        self.assertEqual(edad_cumplida(date(2008, 2, 29), date(2026, 2, 28)), 17)
        self.assertTrue(validar_variante_por_edad(
            instrumento,
            {'fecha_nacimiento': date(2008, 8, 6), 'fecha_aplicacion': fecha_aplicacion},
        )['aplicable'])
        self.assertFalse(validar_variante_por_edad(
            instrumento,
            {'fecha_nacimiento': date(2007, 8, 6), 'fecha_aplicacion': fecha_aplicacion},
        )['aplicable'])
        self.assertFalse(validar_variante_por_edad(
            instrumento,
            {'fecha_nacimiento': date(2015, 8, 6), 'fecha_aplicacion': fecha_aplicacion},
        )['aplicable'])

    def test_scid_adolescente_provisional_requiere_revision_y_valida_edad(self):
        instrumento = self._instrumento(
            'scid-ii-adolescentes',
            CalculadoraInstrumento.Estado.NO_DIAGNOSTICA,
        )
        respuestas = [
            SimpleNamespace(
                pregunta=SimpleNamespace(orden=orden, opciones=[]),
                valor_numerico=Decimal('1') if orden in {1, 105} else Decimal('0'),
            )
            for orden in range(1, 120)
        ]
        resultado = calcular_resultado(
            instrumento,
            respuestas,
            {
                'fecha_nacimiento': date(2011, 1, 1),
                'fecha_aplicacion': date(2026, 8, 6),
            },
        )

        self.assertTrue(resultado['requiere_revision_profesional'])
        self.assertIn('No constituye un diagnóstico', resultado['advertencia_larga'])
        self.assertEqual(
            resultado['detalle']['bloques']['Conductas problemáticas']['estado'],
            'Requiere revisión profesional',
        )
        self.assertTrue(resultado['detalle']['revision_manual_requerida'])
        self.assertIn('Omitir_menor_16', resultado['detalle']['reglas_pendientes'][0])
        self.assertEqual(resultado['detalle']['validacion_edad'], 'válida')
        self.assertEqual(
            resultado['detalle']['trazabilidad_calculadora']['clave'],
            'calc-scid-ii-adolescentes-v1',
        )

    def test_scid_adolescente_sin_fecha_de_nacimiento_no_inventa_edad(self):
        instrumento = self._instrumento(
            'scid-ii-adolescentes',
            CalculadoraInstrumento.Estado.NO_DIAGNOSTICA,
        )
        self.assertIsNone(calcular_resultado(instrumento, []))

    def test_plutchik_adolescente_detecta_criticos_y_calculadora_bloqueada_no_ejecuta(self):
        bloqueado = self._instrumento(
            'ersp-plutchik-adolescentes',
            CalculadoraInstrumento.Estado.BLOQUEADA,
        )
        respuestas = [
            SimpleNamespace(
                pregunta=SimpleNamespace(orden=orden, opciones=[]),
                valor_numerico=Decimal('1') if orden in {1, 13, 15} else Decimal('0'),
            )
            for orden in range(1, 16)
        ]
        contexto = {
            'fecha_nacimiento': date(2010, 8, 6),
            'fecha_aplicacion': date(2026, 8, 6),
        }
        self.assertIsNone(calcular_resultado(bloqueado, respuestas, contexto))
        bloqueado.calculadoras.update(estado=CalculadoraInstrumento.Estado.ORIENTATIVA)

        resultado = calcular_resultado(bloqueado, respuestas, contexto)

        self.assertEqual(resultado['puntaje_total'], Decimal('3'))
        self.assertEqual(
            resultado['detalle']['reactivos_criticos_afirmativos'],
            [13, 15],
        )
        self.assertTrue(resultado['detalle']['existe_respuesta_critica'])
        self.assertTrue(resultado['detalle']['revision_prioritaria'])
        self.assertTrue(resultado['requiere_revision_profesional'])
        self.assertIn('revisión clínica prioritaria', resultado['detalle']['advertencia_prioritaria'])
        self.assertIn('no debe interpretarse como diagnóstico', resultado['advertencia_larga'])


class ImportacionInstrumentoMigrationTests(TransactionTestCase):
    migrate_from = [
        (
            'portafolio',
            (
                '0005_documento_huella_contenido_'
                'instrumento_metadatos_and_more'
            ),
        ),
    ]

    migrate_to = [
        (
            'portafolio',
            (
                '0006_remove_documento_'
                'huella_contenido_and_more'
            ),
        ),
    ]

    def setUp(self):
        super().setUp()

        self.executor = MigrationExecutor(connection)

        self.executor.migrate(
            self.migrate_from
        )

        self.apps_0005 = (
            self.executor.loader
            .project_state(
                self.migrate_from
            )
            .apps
        )

    def tearDown(self):
        self.executor.loader.build_graph()

        self.executor.migrate(
            self.executor.loader.graph.leaf_nodes()
        )

        super().tearDown()

    def _crear(
        self,
        clave,
        con_documento,
        metadatos,
    ):
        Documento = self.apps_0005.get_model(
            'portafolio',
            'Documento',
        )

        InstrumentoHistorico = self.apps_0005.get_model(
            'portafolio',
            'Instrumento',
        )

        documento = (
            Documento.objects.create(
                nombre='Origen ' + clave,
                archivo=clave + '.xlsx',
                huella_contenido='h-' + clave,
            )
            if con_documento
            else None
        )

        return (
            InstrumentoHistorico.objects.create(
                nombre='Instrumento ' + clave,
                clave=clave,
                version='1.0',
                documento_origen=documento,
                metadatos=metadatos,
            ),
            documento,
        )

    def _aplicar_0006(self):
        self.executor = MigrationExecutor(connection)

        self.executor.migrate(
            self.migrate_to
        )

        return (
            self.executor.loader
            .project_state(
                self.migrate_to
            )
            .apps
        )

    def test_transfiere_instrumento_con_documento(self):
        metadatos = {
            'variante': 'adolescentes',
            'edad_min': 12,
            'activo': True,
            'anidado': {
                'valor': None,
            },
        }

        instrumento, documento = self._crear(
            'hist-con-doc',
            True,
            metadatos,
        )

        apps = self._aplicar_0006()

        Importacion = apps.get_model(
            'portafolio',
            'ImportacionInstrumento',
        )

        registro = Importacion.objects.get(
            instrumento_id=instrumento.id
        )

        self.assertEqual(
            (
                registro.instrumento_id,
                registro.documento_id,
                registro.huella_contenido,
                registro.metadatos,
            ),
            (
                instrumento.id,
                documento.id,
                'h-hist-con-doc',
                metadatos,
            ),
        )

        self.assertEqual(
            Importacion.objects.filter(
                instrumento_id=instrumento.id
            ).count(),
            1,
        )

    def test_no_inventa_trazabilidad_sin_documento(self):
        instrumento, _ = self._crear(
            'hist-sin-doc',
            False,
            {
                'origen': 'histórico',
            },
        )

        apps = self._aplicar_0006()

        InstrumentoPosterior = apps.get_model(
            'portafolio',
            'Instrumento',
        )

        Importacion = apps.get_model(
            'portafolio',
            'ImportacionInstrumento',
        )

        self.assertTrue(
            InstrumentoPosterior.objects.filter(
                pk=instrumento.id,
                clave='hist-sin-doc',
            ).exists()
        )

        self.assertFalse(
            Importacion.objects.filter(
                instrumento_id=instrumento.id
            ).exists()
        )

    def test_conjunto_mixto_transfiere_solo_los_dos_con_documento(
        self,
    ):
        con_uno, doc_uno = self._crear(
            'mixto-con-uno',
            True,
            {
                'marca': 'uno',
            },
        )

        con_dos, doc_dos = self._crear(
            'mixto-con-dos',
            True,
            {
                'marca': 'dos',
            },
        )

        sin_uno, _ = self._crear(
            'mixto-sin-uno',
            False,
            {},
        )

        sin_dos, _ = self._crear(
            'mixto-sin-dos',
            False,
            {},
        )

        apps = self._aplicar_0006()

        Importacion = apps.get_model(
            'portafolio',
            'ImportacionInstrumento',
        )

        registros = {
            registro.instrumento_id: registro
            for registro in Importacion.objects.all()
        }

        self.assertEqual(
            len(registros),
            2,
        )

        self.assertEqual(
            (
                registros[con_uno.id].documento_id,
                registros[con_uno.id].metadatos,
            ),
            (
                doc_uno.id,
                {
                    'marca': 'uno',
                },
            ),
        )

        self.assertEqual(
            (
                registros[con_dos.id].documento_id,
                registros[con_dos.id].metadatos,
            ),
            (
                doc_dos.id,
                {
                    'marca': 'dos',
                },
            ),
        )

        self.assertNotIn(
            sin_uno.id,
            registros,
        )

        self.assertNotIn(
            sin_dos.id,
            registros,
        )
